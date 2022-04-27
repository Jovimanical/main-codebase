import openpyxl
from datetime import datetime

# from external.tasks.tasks1 import email_and_sms_helper
from gateway_client.email import email_and_sms_helper


def round_up(data, limit=2):
    if not data:
        return 0
    return round(data, limit)


class PayRoll:

    def __init__(self, document):
        self.file = document

    def get_sheet(self):
        try:
            doc = openpyxl.load_workbook(self.file, data_only=True)
            worksheet = doc.worksheets[-1]
            return worksheet
        except:
            return None

    def parse_column_row(self):
        sheet = self.get_sheet()
        data_start_row = 2
        sheet_title_num = 8
        data_start_col = 9
        data_end_num = sheet.max_row - 6
        data_length = sheet.max_column + 1

        doc_range = range(data_start_row, data_length)
        length_range = range(data_start_col, data_end_num)
        if not sheet:
            return None
        data = [
            {
                sheet.cell(row=sheet_title_num, column=y)
                .value.lower()
                .replace(" ", "_"): sheet.cell(row=x, column=y)
                .value
                for y in doc_range
                if sheet.cell(row=x, column=y).value
                if sheet.cell(row=sheet_title_num, column=y).value
            }
            for x in length_range
        ]
        remove_empty_data = [x for x in data if bool(x)]
        return remove_empty_data

    @classmethod
    def construct_email_context(cls, data):
        date = datetime.today()
        context = dict(
            date=date.strftime("%B %d, %Y"),
            month=date.strftime("%B"),
            year=date.year,
            staff_number=data.get("staff_number"),
            staff_name=data.get("name"),
            basic_pay=f"N{round_up(data.get('basic_payment'))}",
            housing_pay=f"N{round_up(data.get('housing_payment'))}",
            transportation_pay=f"N{round_up(data.get('transportation_payment'))}",
            utility_pay=f"N{round_up(data.get('utility_payment'))}",
            entertainment_pay=f"N{round_up(data.get('entertainment_payment'))}",
            dressing_pay=f"N{round_up(data.get('dressing_payment'))}",
            total_pay=f"N{round_up(data.get('total_payment'))}",
            payee=f"N{round_up(data.get('payee'))}",
            pension=f"N{round_up(data.get('pension_deduction'))}",
            total_deduction=f"N{round_up(data.get('total_deduction'))}",
            net_pay=f"N{round_up(data.get('net_pay'))}",
            position=data.get("position"),
            email=data.get("email"),
        )
        return context

    def send_mail(
        self,
        sms_options=None,
        backend="sendgrid_backend",
        from_mail="Tuteria <automated@tuteria.com>",
        **kwargs,
    ):
        column_row = self.parse_column_row()
        if not column_row:
            return None
        for staff in column_row:
            context = PayRoll.construct_email_context(staff)
            data = dict(
                context=context,
                to=context.get("email"),
                template="payroll",
                title=f"Pay Advice {staff.get('name')}",
                subject=f"Pay Advice {staff.get('name')}",
            )
            email_and_sms_helper(
                data, sms_options=sms_options, backend=backend, from_mail=from_mail
            )
        return True
